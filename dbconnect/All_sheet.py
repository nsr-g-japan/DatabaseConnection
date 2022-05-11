import csv

import openpyxl
from openpyxl import load_workbook

ws = load_workbook('D:\Demotest01.xlsx')
sheets=ws.get_sheet_names()
print(sheets)
sheet_value_arr = []
for a in sheets:
    sheet = ws[a]
    f=open('D:\\totalsheet'+ a +'.csv', 'w')
    c = csv.writer(f)


    for r in sheet.rows:
        c.writerow([cell.value for cell in r])

        #sheet_value_arr.append([cell.value for cell in row])
        #for r in sheet.rows:
            #c.writerow([cell.value for cell in row])

f.close()